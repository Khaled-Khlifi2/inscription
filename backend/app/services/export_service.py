"""
Service d'export d'étudiants — flexible et configurable.

Architecture
────────────
1. **FIELD_CATALOG** : description centralisée de tous les champs exportables
   (clé, libellé par défaut, groupe, type, getter).
2. **PRESETS** : ensembles prédéfinis de colonnes (SALIMA, Minimal, Coordonnées…).
3. **export_custom()** : génère xlsx ou csv à partir d'une sélection de colonnes
   avec libellés personnalisés.
4. **export_to_xlsx()** : raccourci legacy → preset SALIMA en xlsx.
"""
from __future__ import annotations

from io import BytesIO, StringIO
from typing import Any, Callable, Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.models import Etudiant, Inscription, Niveau

ANNEE_EN_COURS = "2025/2026"


# ──────────────────────────────────────────────────────────────────
#  Catalogue des champs exportables
# ──────────────────────────────────────────────────────────────────
# Chaque entrée : (key, default_label, group, type, getter)
#   - type : 'text' | 'arabic' | 'date' | 'datetime' | 'bool' | 'int'
#   - getter : callable(Etudiant) -> Any  (None si direct attribute)

def _g(attr: str) -> Callable[[Etudiant], Any]:
    return lambda e, a=attr: getattr(e, a, None)


def _bool_oui_non(v: Any) -> str:
    if v is None:
        return ""
    return "Oui" if bool(v) else "Non"


def _format_value(value: Any, type_: str) -> Any:
    """Normalise une valeur pour l'export (xlsx + csv)."""
    if value is None:
        return ""
    if type_ == "bool":
        return _bool_oui_non(value)
    if type_ == "datetime":
        try:
            return value.strftime("%Y-%m-%d %H:%M")
        except Exception:
            return str(value)
    if type_ == "date":
        return str(value)
    return value


FIELD_CATALOG: list[dict[str, Any]] = [
    # ─── Identité administrative ───
    {"key": "mat_cin",         "label": "MAT_CIN",         "group": "Identité administrative", "type": "text"},
    {"key": "num_inscription", "label": "NUM_INSCRIPTION", "group": "Identité administrative", "type": "text"},

    # ─── Identité civile (FR) ───
    {"key": "nom_fr",          "label": "NOM_FR",          "group": "Identité civile",         "type": "text"},
    {"key": "prenom_fr",       "label": "PRENOM_FR",       "group": "Identité civile",         "type": "text"},
    {"key": "sexe",            "label": "SEXE",            "group": "Identité civile",         "type": "text"},
    {"key": "situation_familiale", "label": "SITUATION_FAMILIALE", "group": "Identité civile", "type": "text"},

    # ─── Identité civile (AR) ───
    {"key": "nom_ar",          "label": "NOM_AR",          "group": "Identité civile (AR)",    "type": "arabic"},
    {"key": "prenom_ar",       "label": "PRENOM_AR",       "group": "Identité civile (AR)",    "type": "arabic"},

    # ─── Naissance ───
    {"key": "date_naissance",  "label": "DATE_NAISSANCE",  "group": "Naissance",               "type": "date"},
    {"key": "lieu_naiss_fr",   "label": "LIEU_NAISSANCE_FR", "group": "Naissance",             "type": "text"},
    {"key": "lieu_naiss_ar",   "label": "LIEU_NAISSANCE_AR", "group": "Naissance",             "type": "arabic"},

    # ─── Coordonnées ───
    {"key": "email",            "label": "EMAIL",            "group": "Coordonnées",            "type": "text"},
    {"key": "telephone_portable","label": "TEL_PORTABLE",    "group": "Coordonnées",            "type": "text"},
    {"key": "telephone_fixe",   "label": "TEL_FIXE",         "group": "Coordonnées",            "type": "text"},
    {"key": "adresse_fr",       "label": "ADRESSE_FR",       "group": "Coordonnées",            "type": "text"},
    {"key": "adresse_ar",       "label": "ADRESSE_AR",       "group": "Coordonnées",            "type": "arabic"},

    # ─── Données administratives ───
    {"key": "code_gouvernorat", "label": "CODE_GOUVERNORAT", "group": "Administratif",          "type": "text"},
    {"key": "code_type_bac",    "label": "CODE_TYPE_BAC",    "group": "Administratif",          "type": "text"},
    {"key": "num_cnss",         "label": "NUM_CNSS",         "group": "Administratif",          "type": "text"},
    {"key": "passeport",        "label": "PASSEPORT",        "group": "Administratif",          "type": "text"},
    {"key": "statut",           "label": "STATUT",           "group": "Administratif",          "type": "text"},

    # ─── Cursus académique ───
    {"key": "cfil",             "label": "CFIL",             "group": "Cursus académique",      "type": "text"},
    {"key": "lib_filiere",      "label": "LIB_FILIERE",      "group": "Cursus académique",      "type": "text"},
    {"key": "lib_filiere_ar",   "label": "LIB_FILIERE_AR",   "group": "Cursus académique",      "type": "arabic"},
    {"key": "niveau_libelle",   "label": "NIVEAU",           "group": "Cursus académique",      "type": "text",
     "getter": lambda e: e.niveau.libelle if e.niveau else None},
    {"key": "niveau_code",      "label": "NIVEAU_CODE",      "group": "Cursus académique",      "type": "text",
     "getter": lambda e: e.niveau.code if e.niveau else None},

    # ─── Statut & dates système ───
    {"key": "is_active",                "label": "COMPTE_ACTIF",        "group": "Statut & dates", "type": "bool"},
    {"key": "email_verified",           "label": "EMAIL_VERIFIE",       "group": "Statut & dates", "type": "bool"},
    {"key": "is_inscription_complete",  "label": "DOSSIER_COMPLET",     "group": "Statut & dates", "type": "bool"},
    {"key": "completed_at",             "label": "DATE_SOUMISSION",     "group": "Statut & dates", "type": "datetime"},
    {"key": "email_verified_at",        "label": "DATE_VERIF_EMAIL",    "group": "Statut & dates", "type": "datetime"},
    {"key": "created_at",               "label": "DATE_CREATION",       "group": "Statut & dates", "type": "datetime"},
    {"key": "updated_at",               "label": "DATE_MAJ",            "group": "Statut & dates", "type": "datetime"},
]

CATALOG_BY_KEY: dict[str, dict[str, Any]] = {f["key"]: f for f in FIELD_CATALOG}


# ──────────────────────────────────────────────────────────────────
#  Presets de colonnes
# ──────────────────────────────────────────────────────────────────
PRESETS: dict[str, dict[str, Any]] = {
    "salima": {
        "label": "SALIMA — Format Ministère",
        "description": "Format officiel pour le Ministère de l'Enseignement Supérieur",
        "columns": [
            "num_inscription", "mat_cin",
            "nom_ar", "prenom_ar", "nom_fr", "prenom_fr",
            "sexe", "situation_familiale",
            "date_naissance", "lieu_naiss_ar",
            "code_gouvernorat", "code_type_bac",
            "num_cnss", "passeport",
            "cfil", "lib_filiere",
        ],
    },
    "minimal": {
        "label": "Minimal — Identification",
        "description": "Matricule, nom complet et filière",
        "columns": [
            "mat_cin", "nom_fr", "prenom_fr",
            "niveau_libelle", "lib_filiere",
        ],
    },
    "contact": {
        "label": "Coordonnées — Communication",
        "description": "Pour envoi de mails / SMS",
        "columns": [
            "mat_cin", "nom_fr", "prenom_fr",
            "email", "telephone_portable", "telephone_fixe",
            "adresse_fr",
        ],
    },
    "complete": {
        "label": "Complet — Toutes les données",
        "description": "Toutes les colonnes disponibles",
        "columns": [f["key"] for f in FIELD_CATALOG],
    },
}


# ──────────────────────────────────────────────────────────────────
#  Service
# ──────────────────────────────────────────────────────────────────
class ExportService:

    # --------- Catalogue ---------
    @staticmethod
    def get_catalog() -> dict[str, Any]:
        """
        Retourne le catalogue de champs exportables.
        Les modèles d'export sont gérés côté frontend (localStorage).
        """
        return {
            "fields": [
                {"key": f["key"], "label": f["label"], "group": f["group"], "type": f["type"]}
                for f in FIELD_CATALOG
            ],
        }

    # --------- Récupération données ---------
    @staticmethod
    async def _fetch_etudiants(
        db: AsyncSession,
        *,
        cfil: Optional[str] = None,
        niveau_id: Optional[int] = None,
        inscription_only: bool = False,
    ) -> list[Etudiant]:
        q = select(Etudiant).options(selectinload(Etudiant.niveau)).where(Etudiant.is_active == True)  # noqa: E712
        if cfil:
            q = q.where(Etudiant.cfil == cfil)
        if niveau_id:
            q = q.where(Etudiant.niveau_id == niveau_id)
        if inscription_only:
            # uniquement les étudiants ayant une inscription "validee" pour l'année en cours
            sub = (
                select(Inscription.etudiant_id)
                .where(
                    Inscription.annee_universitaire == ANNEE_EN_COURS,
                    Inscription.statut == "validee",
                )
            )
            q = q.where(Etudiant.id.in_(sub))
        q = q.order_by(Etudiant.nom_fr, Etudiant.prenom_fr)
        result = await db.execute(q)
        return list(result.scalars().all())

    # --------- Extraction valeur d'un champ ---------
    @staticmethod
    def _extract(e: Etudiant, key: str) -> Any:
        meta = CATALOG_BY_KEY.get(key)
        if not meta:
            return ""
        getter = meta.get("getter") or _g(key)
        return _format_value(getter(e), meta["type"])

    # --------- Build des lignes ---------
    @staticmethod
    def _build_rows(etudiants: list[Etudiant], keys: list[str]) -> list[list[Any]]:
        return [[ExportService._extract(e, k) for k in keys] for e in etudiants]

    # --------- Export XLSX ---------
    @staticmethod
    def _to_xlsx(headers: list[str], rows: list[list[Any]], sheet_title: str = "Étudiants") -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title[:31]

        # En-têtes stylisées
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="1F2937")
        center = Alignment(horizontal="center", vertical="center")

        for col_idx, label in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for r_idx, row in enumerate(rows, start=2):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        # Largeurs auto
        for col_idx, label in enumerate(headers, start=1):
            max_len = max(
                [len(str(label))] + [len(str(r[col_idx - 1])) for r in rows[:200]]
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

        ws.freeze_panes = "A2"

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # --------- Export CSV ---------
    @staticmethod
    def _to_csv(headers: list[str], rows: list[list[Any]]) -> bytes:
        import csv
        buf = StringIO()
        writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(["" if v is None else v for v in row])
        # BOM utf-8 pour qu'Excel ouvre correctement les caractères arabes/accents
        return ("\ufeff" + buf.getvalue()).encode("utf-8")

    # ══════════════════════════════════════════════════════════════
    #  API publique
    # ══════════════════════════════════════════════════════════════
    @staticmethod
    async def export_custom(
        db: AsyncSession,
        *,
        columns: list[dict[str, str]],          # [{"key":..., "label":...}]
        fmt: str = "xlsx",                      # "xlsx" | "csv"
        cfil: Optional[str] = None,
        niveau_id: Optional[int] = None,
        inscription_only: bool = False,
        sheet_title: str = "Étudiants",
    ) -> tuple[bytes, str, str]:
        """
        Génère un fichier d'export selon les colonnes choisies.

        Returns (bytes, mime_type, extension)
        """
        # Filtrer & valider les colonnes demandées
        valid = [c for c in columns if c.get("key") in CATALOG_BY_KEY]
        if not valid:
            raise ValueError("Aucune colonne valide sélectionnée")

        keys = [c["key"] for c in valid]
        labels = [c.get("label") or CATALOG_BY_KEY[c["key"]]["label"] for c in valid]

        etudiants = await ExportService._fetch_etudiants(
            db, cfil=cfil, niveau_id=niveau_id, inscription_only=inscription_only,
        )
        rows = ExportService._build_rows(etudiants, keys)

        if fmt == "csv":
            return (
                ExportService._to_csv(labels, rows),
                "text/csv; charset=utf-8",
                "csv",
            )

        return (
            ExportService._to_xlsx(labels, rows, sheet_title=sheet_title),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xlsx",
        )

    # --------- Legacy : preset SALIMA en xlsx (compat existante) ---------
    @staticmethod
    async def export_to_xlsx(
        db: AsyncSession,
        *,
        cfil: Optional[str] = None,
        niveau_id: Optional[int] = None,
        inscription_only: bool = True,
    ) -> bytes:
        preset_keys = PRESETS["salima"]["columns"]
        columns = [{"key": k, "label": CATALOG_BY_KEY[k]["label"]} for k in preset_keys]
        data, _, _ = await ExportService.export_custom(
            db,
            columns=columns,
            fmt="xlsx",
            cfil=cfil,
            niveau_id=niveau_id,
            inscription_only=inscription_only,
            sheet_title="SALIMA",
        )
        return data
